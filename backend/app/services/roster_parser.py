"""花名册 Excel 文件解析服务。"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import uuid
from datetime import date, datetime, timedelta
from typing import Any, Optional

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

from ..exceptions import ValidationError
from .normalize import normalize_name

logger = logging.getLogger(__name__)

_EXCEL_DATE_EPOCH = date(1899, 12, 30)
"""Excel 日期序列号纪元（处理 1900 闰年 Bug）。"""

_MAX_DATE_SERIAL = 2958465
"""Excel 最大有效日期序列号（9999-12-31）。"""


def parse_excel(
    file_path: str,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> dict:
    """解析花名册 Excel 文件（.xlsx / .xlsm）。

    打开工作簿两次：一次读取公式（data_only=False），
    一次读取缓存值（data_only=True），以同时获取公式文本与计算结果。
    自动检测工作表名称和表头行。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 工作表名称。为 None 时自动检测。
        header_row: 预设表头行号（1-based），仅在自动检测失败时使用。

    Returns:
        dict: 包含以下键：
            - sheet_names (list[str]): 所有工作表名称
            - selected_sheet (str): 选中的工作表名称
            - header_row (int): 实际使用的表头行号
            - headers (list[str]): 规范化后的表头列表
            - rows (list[dict]): 数据行，每行以规范化表头为键
            - raw_data (list[dict]): 原始单元格值行列表
            - formulas (dict[str, str]): 公式字典（cell_ref -> formula）
            - cached_values (dict): 公式缓存值字典
            - images (dict[str, str]): 单元格图片路径字典
            - detected_header_row (int): 自动检测到的表头行号（0 表示未找到）

    Raises:
        ValidationError: 文件不存在、格式不支持或工作表不存在。
    """
    _validate_file(file_path)

    wb_formula = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)

    try:
        sheet_names = wb_formula.sheetnames

        if sheet_name is None:
            sheet_name = _detect_sheet(sheet_names)
        elif sheet_name not in sheet_names:
            raise ValidationError(
                f"工作表 '{sheet_name}' 不存在，可选: {', '.join(sheet_names)}"
            )

        ws_formula = wb_formula[sheet_name]
        ws_data = wb_data[sheet_name]

        detected_header_row = detect_header_row(file_path, sheet_name)
        actual_header_row = detected_header_row if detected_header_row > 0 else header_row

        headers = _parse_headers(ws_formula, actual_header_row)
        rows, raw_data, formulas, cached_values = _parse_rows(
            ws_formula, ws_data, headers, actual_header_row,
        )

        # 提取图片（仅对 .xlsm 执行）
        images: dict[str, str] = {}
        if file_path.lower().endswith(".xlsm"):
            output_dir = os.path.join(os.path.dirname(file_path), "_images")
            os.makedirs(output_dir, exist_ok=True)
            images = _extract_images(file_path, sheet_name, output_dir)

        return {
            "sheet_names": sheet_names,
            "selected_sheet": sheet_name,
            "header_row": actual_header_row,
            "headers": headers,
            "rows": rows,
            "raw_data": raw_data,
            "formulas": formulas,
            "cached_values": cached_values,
            "images": images,
            "detected_header_row": detected_header_row,
        }
    finally:
        wb_formula.close()
        wb_data.close()


def detect_sheet_names(file_path: str) -> list[str]:
    """获取工作簿的所有工作表名称。

    Args:
        file_path: Excel 文件路径。

    Returns:
        list[str]: 工作表名称列表。

    Raises:
        ValidationError: 文件无效。
    """
    _validate_file(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def detect_header_row(file_path: str, sheet_name: str) -> int:
    """自动检测表头行号。

    在前 10 行中查找包含"姓名"或"name"（不区分大小写）的单元格。
    返回匹配到的第一个行号（1-based）。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 工作表名称。

    Returns:
        int: 表头行号（1-based），未找到时返回 0。
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return 0

            ws = wb[sheet_name]
            for row_idx, row in enumerate(
                ws.iter_rows(max_row=10, values_only=True), start=1
            ):
                for cell_value in row:
                    if cell_value is not None:
                        text = str(cell_value).strip()
                        if re.search(r"(姓名|name)", text, re.IGNORECASE):
                            return row_idx
            return 0
        finally:
            wb.close()
    except Exception:
        logger.warning(
            "检测表头行失败: file=%s sheet=%s", file_path, sheet_name, exc_info=True
        )
        return 0


def save_uploaded_file(uploaded_file, storage_dir: str) -> dict:
    """保存上传的文件到存储目录。

    使用 UUID 文件名避免冲突，同时记录原始文件名、
    SHA256 哈希值和文件大小。

    Args:
        uploaded_file: 上传文件对象，须有 .filename 属性和 .read() 方法。
        storage_dir: 存储目录路径。

    Returns:
        dict: 包含 stored_path, original_name, sha256, size 四个键。
    """
    os.makedirs(storage_dir, exist_ok=True)

    original_name = _sanitize_filename(
        getattr(uploaded_file, "filename", "unknown.xlsx")
    )
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(original_name)[1] or ".xlsx"
    stored_name = f"{file_id}{ext}"
    stored_path = os.path.abspath(os.path.join(storage_dir, stored_name))

    content = uploaded_file.read()
    sha256_hash = hashlib.sha256(content).hexdigest()
    file_size = len(content)

    with open(stored_path, "wb") as f:
        f.write(content)

    return {
        "stored_path": stored_path,
        "original_name": original_name,
        "sha256": sha256_hash,
        "size": file_size,
    }


# ── Private API ─────────────────────────────────────────────────────


def _normalize_header(header_text: str | None) -> str:
    """规范化表头文本。

    处理规则：
    1. None 或空输入 → 返回空字符串。
    2. 去除首尾空白。
    3. 清除特殊字符，仅保留汉字、字母、数字、下划线和常见括号。

    Args:
        header_text: 原始表头文本。

    Returns:
        str: 规范化后的表头文本。空列头返回空字符串。
    """
    if header_text is None:
        return ""
    text = str(header_text).strip()
    if not text:
        return ""
    text = re.sub(r"[^\w一-鿿\-（）()【\]\[\]]", "", text)
    return text


def _parse_cell_date(
    cell_value: Any,
    data_only_wb: openpyxl.Workbook | None = None,
) -> date | None:
    """解析单元格值为 date 对象。

    处理以下类型：
    - datetime.datetime → 提取 date 部分
    - datetime.date → 直接返回
    - int / float → 按 Excel 序列号转换
    - str → 尝试多种日期字符串格式

    所有日期返回 date 对象，不包含时区信息。

    Args:
        cell_value: 单元格原始值。
        data_only_wb: 可选，用于扩展日期识别逻辑的 Workbook 引用。

    Returns:
        date | None: 解析成功返回 date 对象，否则返回 None。
    """
    if cell_value is None:
        return None

    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, date):
        return cell_value

    if isinstance(cell_value, (int, float)):
        return _excel_serial_to_date(cell_value)

    if isinstance(cell_value, str):
        return _parse_date_string(cell_value.strip())

    return None


def _parse_date_string(date_str: str) -> date | None:
    """解析日期字符串为 date 对象。

    支持的格式：
    - 2026-07-16
    - 2026/7/16
    - 2026.7.16
    - 20260716
    - 2026年7月16日

    对于仅有月日无年份的模糊日期，返回 None。

    Args:
        date_str: 日期字符串。

    Returns:
        date | None: 解析成功返回 date 对象，否则返回 None。
    """
    if not date_str:
        return None

    # "2026年7月16日" / "2026年07月16日"
    m = re.match(
        r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?", date_str
    )
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None

    # "2026-07-16" / "2026/7/16" / "2026.7.16"
    m = re.match(
        r"(\d{4})\s*[-/.]\s*(\d{1,2})\s*[-/.]\s*(\d{1,2})", date_str
    )
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None

    # "20260716"（8 位连续数字）
    m = re.match(r"(\d{4})(\d{2})(\d{2})$", date_str)
    if m:
        year, month, day = (
            int(m.group(1)),
            int(m.group(2)),
            int(m.group(3)),
        )
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                return date(year, month, day)
            except ValueError:
                return None

    # 仅有月日无年份 → 不明确，返回 None
    return None


def _clean_text(value: Any) -> str:
    """清洗文本值。

    步骤：
    1. 转换为字符串。
    2. 去除首尾空白。
    3. 全角字母、数字、空格 → 半角。
    4. 连续空白 → 单个空格。

    Args:
        value: 原始单元格值。

    Returns:
        str: 清洗后的文本。
    """
    if value is None:
        return ""
    text = str(value).strip()
    text = _fullwidth_to_halfwidth(text)
    text = _collapse_whitespace(text)
    return text


def _extract_images(
    file_path: str,
    sheet_name: str,
    output_dir: str,
) -> dict[str, str]:
    """提取工作表中嵌入单元格的图片。

    仅对 .xlsm 文件有效。xlsx 文件通常不将图片嵌入单元格。
    openpyxl 的图片提取能力有限，使用 try/except 兜底。

    Args:
        file_path: Excel 文件路径。
        sheet_name: 工作表名称。
        output_dir: 图片输出目录。

    Returns:
        dict[str, str]: cell_ref -> 图片保存路径的字典。
    """
    images: dict[str, str] = {}
    try:
        wb = openpyxl.load_workbook(file_path)
        try:
            if sheet_name not in wb.sheetnames:
                return images
            ws = wb[sheet_name]

            drawings = getattr(ws, "_images", None)
            if not drawings:
                return images

            for drawing in drawings:
                if not isinstance(drawing, OpenpyxlImage):
                    continue
                try:
                    cell_ref = _resolve_image_anchor(drawing)
                    if cell_ref is None:
                        continue

                    img_bytes = _get_image_bytes(drawing)
                    if not img_bytes:
                        continue

                    img_filename = f"{uuid.uuid4()}.png"
                    img_path = os.path.abspath(
                        os.path.join(output_dir, img_filename)
                    )
                    with open(img_path, "wb") as f:
                        f.write(img_bytes)
                    images[cell_ref] = img_path
                except Exception:
                    logger.debug(
                        "跳过图片提取: file=%s drawing=%s",
                        file_path,
                        drawing,
                        exc_info=True,
                    )
                    continue
        finally:
            wb.close()
    except Exception:
        logger.warning(
            "提取图片失败: file=%s sheet=%s", file_path, sheet_name, exc_info=True
        )

    return images


def _fullwidth_to_halfwidth(text: str) -> str:
    """将全角字母、数字、空格转换为半角。

    全角字符范围 FF01-FF5E 映射到半角 21-7E。
    全角空格 U+3000 映射到半角空格 U+0020。

    Args:
        text: 输入文本。

    Returns:
        str: 转换后的文本。
    """
    result: list[str] = []
    for char in text:
        code = ord(char)
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        elif code == 0x3000:
            result.append(" ")
        else:
            result.append(char)
    return "".join(result)


def _collapse_whitespace(text: str) -> str:
    """将连续空白字符压缩为单个空格。

    Args:
        text: 输入文本。

    Returns:
        str: 压缩后的文本。
    """
    return re.sub(r"\s+", " ", text)


def _validate_file(file_path: str) -> None:
    """验证 Excel 文件路径有效性。

    Args:
        file_path: 文件路径。

    Raises:
        ValidationError: 文件不存在或格式不支持。
    """
    if not os.path.isfile(file_path):
        raise ValidationError(f"文件不存在: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        raise ValidationError("请另存为 xlsx 后再上传")
    if ext not in (".xlsx", ".xlsm"):
        raise ValidationError(
            f"不支持的文件格式 '{ext}'，仅支持 .xlsx 和 .xlsm"
        )


# ── Internal helpers（辅助公共 API，不对外暴露）──────────────────────


def _detect_sheet(sheet_names: list[str]) -> str:
    """自动检测最佳工作表。

    优先级：
    1. "员工花名册-总表"
    2. 名称包含"花名册"的工作表
    3. 第一个工作表

    Args:
        sheet_names: 所有工作表名称列表。

    Returns:
        str: 选中的工作表名称。
    """
    preferred = "员工花名册-总表"
    if preferred in sheet_names:
        return preferred
    for name in sheet_names:
        if "花名册" in name:
            return name
    return sheet_names[0]


def _parse_headers(ws: Any, header_row: int) -> list[str]:
    """解析并规范化表头行。

    Args:
        ws: openpyxl Worksheet 对象。
        header_row: 表头行号（1-based）。

    Returns:
        list[str]: 规范化且唯一的表头列表。
    """
    raw: list[str] = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        raw.append(_normalize_header(cell.value))
    return _build_unique_headers(raw)


def _build_unique_headers(headers: list[str]) -> list[str]:
    """确保表头名称唯一，重复项添加 _N 后缀。

    例如 ['姓名', '姓名'] → ['姓名', '姓名_2']。

    Args:
        headers: 原始表头列表（可含重复值）。

    Returns:
        list[str]: 保证无重复的表头列表。
    """
    seen: dict[str, int] = {}
    result: list[str] = []
    for h in headers:
        if h not in seen:
            seen[h] = 1
            result.append(h)
        else:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
    return result


def _parse_cell_value(value: Any) -> Any:
    """将单元格值转换为 Python 原生类型。

    - datetime → date（时区无关）
    - str → 经 _clean_text 清洗
    - int / float / bool / None → 保持原样
    - 其他 → str(value)

    Args:
        value: 原始单元格值。

    Returns:
        转换后的值。
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return _clean_text(value)
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _parse_rows(
    ws_formula: Any,
    ws_data: Any,
    headers: list[str],
    header_row: int,
) -> tuple[list[dict], list[dict], dict[str, str], dict[str, Any]]:
    """解析表头行之后的所有数据行。

    检测公式单元格（值以 '=' 开头），将公式文本存入 formulas 字典，
    缓存值存入 cached_values 字典。无任何实质内容的空行被自动跳过。

    Args:
        ws_formula: data_only=False 模式的工作表（含公式字符串）。
        ws_data: data_only=True 模式的工作表（含缓存计算值）。
        headers: 规范化表头列表。
        header_row: 表头行号。

    Returns:
        (rows, raw_data, formulas, cached_values) 四元组。
    """
    rows: list[dict] = []
    raw_data: list[dict] = []
    formulas: dict[str, str] = {}
    cached_values: dict[str, Any] = {}
    header_count = len(headers)

    for row_idx in range(header_row + 1, ws_formula.max_row + 1):
        row_dict: dict[str, Any] = {}
        row_raw: dict[str, Any] = {}
        has_data = False

        for col_idx in range(1, header_count + 1):
            header = headers[col_idx - 1]
            cell_f = ws_formula.cell(row=row_idx, column=col_idx)
            cell_d = ws_data.cell(row=row_idx, column=col_idx)

            value_f = cell_f.value
            is_formula = isinstance(value_f, str) and value_f.startswith("=")

            if is_formula:
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                formulas[cell_ref] = value_f
                raw_val = cell_d.value
                cached_values[cell_ref] = raw_val
                value = _parse_cell_value(raw_val)
            else:
                value = _parse_cell_value(value_f)

            row_dict[header] = value
            row_raw[get_column_letter(col_idx)] = value_f

            if value is not None and (not isinstance(value, str) or value.strip()):
                has_data = True

        if has_data:
            rows.append(row_dict)
            raw_data.append(row_raw)

    return rows, raw_data, formulas, cached_values


def _sanitize_filename(filename: str) -> str:
    """净化文件名，防止路径遍历攻击。

    去除路径部分（保留 basename）和危险字符。

    Args:
        filename: 原始文件名。

    Returns:
        str: 净化后的安全文件名。
    """
    filename = os.path.basename(filename)
    filename = filename.replace("\0", "")
    filename = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", filename)
    if not filename:
        filename = "unknown.xlsx"
    return filename


def _excel_serial_to_date(serial: int | float) -> date | None:
    """将 Excel 日期序列号转换为 date 对象。

    处理 Excel 1900 闰年 Bug（序列号 60 对应不存在的 1900-02-29）。

    Args:
        serial: Excel 日期序列号。

    Returns:
        date | None: 转换成功返回 date，否则返回 None。
    """
    if not isinstance(serial, (int, float)):
        return None
    if serial < 1 or serial > _MAX_DATE_SERIAL:
        return None
    try:
        return _EXCEL_DATE_EPOCH + timedelta(days=int(serial))
    except (OverflowError, ValueError):
        return None


def _resolve_image_anchor(drawing: OpenpyxlImage) -> str | None:
    """从图片对象的锚点中解析出关联的单元格引用。

    openpyxl 使用 0-based 行列索引，Excel 使用 1-based 行列索引。

    Args:
        drawing: openpyxl Image 对象。

    Returns:
        str | None: 单元格引用（如 "B3"），无法解析返回 None。
    """
    try:
        anchor = drawing.anchor
        if hasattr(anchor, "_from"):
            col_idx = anchor._from.col
            row_idx = anchor._from.row + 1  # openpyxl 锚点是 0-based
            if col_idx >= 0 and row_idx > 0:
                return f"{get_column_letter(col_idx + 1)}{row_idx}"
    except Exception:
        pass
    return None


def _get_image_bytes(drawing: OpenpyxlImage) -> bytes | None:
    """从 drawing 中提取图片字节数据。

    Args:
        drawing: openpyxl Image 对象。

    Returns:
        bytes | None: 图片字节数据，提取失败返回 None。
    """
    try:
        if hasattr(drawing, "_data") and callable(drawing._data):
            data = drawing._data()
            if isinstance(data, bytes):
                return data
            if hasattr(data, "read"):
                return data.read()
        if hasattr(drawing, "data"):
            data = drawing.data
            if isinstance(data, bytes):
                return data
            if callable(data):
                data = data()
                if isinstance(data, bytes):
                    return data
                if hasattr(data, "read"):
                    return data.read()
        if hasattr(drawing, "ref"):
            # 通过 ref 间接读取，兜底路径
            pass
    except Exception:
        pass
    return None
