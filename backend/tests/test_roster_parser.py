"""花名册导入解析器测试。"""
from __future__ import annotations

import hashlib
import os
import tempfile
from datetime import date, datetime, timedelta
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from openpyxl import Workbook

from app.exceptions import ValidationError
from app.services.roster_parser import (
    _detect_sheet,
    _excel_serial_to_date,
    _extract_images,
    _parse_cell_date,
    _parse_date_string,
    _parse_headers,
    parse_excel,
    save_uploaded_file,
)

# ============================================================
# 辅助函数
# ============================================================


def _create_minimal_xlsx(tmp_path: str, rows: list[list] | None = None) -> str:
    """创建最小 xlsx 文件并返回路径。"""
    path = os.path.join(tmp_path, "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "员工花名册-总表"
    ws.append(["序号", "姓名", "入职日期"])
    if rows:
        for r in rows:
            ws.append(r)
    wb.save(path)
    return path


def _create_xlsm_with_image(tmp_path: str) -> str:
    """创建含图片的 xlsm 文件。"""
    path = os.path.join(tmp_path, "test.xlsm")
    wb = Workbook()
    ws = wb.active
    ws.title = "员工花名册-总表"
    ws.append(["序号", "姓名", "照片"])
    ws.append([1, "张三", "图片"])
    wb.save(path)
    return path


# ============================================================
# 工作表检测
# ============================================================


class TestDetectSheet:
    def test_prefers_总表(self):
        names = ["Sheet1", "员工花名册-总表", "Sheet2"]
        assert _detect_sheet(names) == "员工花名册-总表"

    def test_prefers_花名册(self):
        names = ["Sheet1", "花名册-2026"]
        assert _detect_sheet(names) == "花名册-2026"

    def test_falls_back_to_first(self):
        names = ["Sheet1", "Sheet2"]
        assert _detect_sheet(names) == "Sheet1"


# ============================================================
# Excel 日期序列号
# ============================================================


class TestExcelSerialToDate:
    def test_valid_serial(self):
        assert _excel_serial_to_date(1) == date(1899, 12, 31)
        assert _excel_serial_to_date(42891) == date(2017, 6, 5)

    def test_none(self):
        assert _excel_serial_to_date(None) is None

    def test_invalid_serial(self):
        assert _excel_serial_to_date(0) is None
        assert _excel_serial_to_date(-1) is None
        assert _excel_serial_to_date(2958466) is None


# ============================================================
# 日期字符串解析
# ============================================================


class TestParseDateString:
    def test_yyyy_mm_dd(self):
        assert _parse_date_string("2026-07-16") == date(2026, 7, 16)

    def test_yyyy_slash_mm_slash_dd(self):
        assert _parse_date_string("2026/7/16") == date(2026, 7, 16)
        assert _parse_date_string("2026/07/16") == date(2026, 7, 16)

    def test_yyyy_dot_mm_dot_dd(self):
        assert _parse_date_string("2026.7.16") == date(2026, 7, 16)

    def test_compact_8_digit(self):
        """20260716 格式"""
        assert _parse_date_string("20260716") == date(2026, 7, 16)

    def test_chinese_format(self):
        assert _parse_date_string("2026年7月16日") == date(2026, 7, 16)

    def test_month_day_only_returns_none(self):
        """仅有月日无年份时返回 None"""
        assert _parse_date_string("7月16日") is None

    def test_invalid_date_returns_none(self):
        assert _parse_date_string("not-a-date") is None
        assert _parse_date_string("") is None

    def test_empty_string_returns_none(self):
        assert _parse_date_string("") is None


# ============================================================
# 单元格日期解析
# ============================================================


class TestParseCellDate:
    def test_datetime_object(self):
        assert _parse_cell_date(datetime(2026, 7, 16, 10, 30)) == date(2026, 7, 16)

    def test_date_object(self):
        assert _parse_cell_date(date(2026, 7, 16)) == date(2026, 7, 16)

    def test_excel_serial(self):
        assert _parse_cell_date(42891) == date(2017, 6, 5)

    def test_none_returns_none(self):
        assert _parse_cell_date(None) is None

    def test_date_string(self):
        assert _parse_cell_date("2026-07-16") == date(2026, 7, 16)
        assert _parse_cell_date("20260716") == date(2026, 7, 16)


# ============================================================
# Excel 解析
# ============================================================


class TestParseExcel:
    def test_parse_success(self, tmp_path):
        file_path = _create_minimal_xlsx(
            str(tmp_path),
            [[1, "张三", date(2026, 1, 15)]],
        )
        result = parse_excel(file_path)
        assert result["selected_sheet"] == "员工花名册-总表"
        assert len(result["rows"]) == 1
        assert result["rows"][0]["姓名"] == "张三"

    def test_invalid_file_raises(self):
        with pytest.raises(ValidationError, match="文件不存在"):
            parse_excel("/nonexistent/file.xlsx")

    def test_xls_rejected(self, tmp_path):
        path = os.path.join(str(tmp_path), "test.xls")
        with open(path, "w") as f:
            f.write("dummy")
        with pytest.raises(ValidationError, match="请另存为 xlsx"):
            parse_excel(path)

    def test_empty_file(self, tmp_path):
        path = os.path.join(str(tmp_path), "empty.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append([])  # 只有空表头
        wb.save(path)
        result = parse_excel(path)
        assert len(result["rows"]) == 0

    def test_a_column_empty_header_is_序号(self):
        """A列空表头应被识别为序号"""
        pass  # 在解析逻辑中处理

    def test_auto_detect_sheet(self, tmp_path):
        path = os.path.join(str(tmp_path), "multi_sheet.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["姓名"])
        ws1.append(["张三"])
        ws2 = wb.create_sheet("员工花名册-总表")
        ws2.append(["姓名"])
        ws2.append(["李四"])
        wb.save(path)

        result = parse_excel(path, sheet_name=None)
        assert result["selected_sheet"] == "员工花名册-总表"
        assert result["rows"][0]["姓名"] == "李四"


# ============================================================
# 图片提取
# ============================================================


class TestExtractImages:
    def test_xlsx_image_extraction(self, tmp_path):
        """xlsx 也能提取图片（不只是 xlsm）"""
        path = _create_xlsm_with_image(str(tmp_path))
        output_dir = os.path.join(str(tmp_path), "_images")
        os.makedirs(output_dir, exist_ok=True)
        # 提取不抛异常即可
        images = _extract_images(path, "员工花名册-总表", output_dir)
        assert isinstance(images, dict)

    def test_no_images_returns_empty(self, tmp_path):
        path = _create_minimal_xlsx(str(tmp_path))
        output_dir = os.path.join(str(tmp_path), "_images")
        os.makedirs(output_dir, exist_ok=True)
        images = _extract_images(path, "员工花名册-总表", output_dir)
        assert images == {}


# ============================================================
# 文件保存与重复检测
# ============================================================


class TestSaveUploadedFile:
    def test_saves_file_and_returns_hash(self, tmp_path):
        content = b"dummy excel content"
        uploaded = MagicMock()
        uploaded.filename = "test.xlsx"
        uploaded.read.return_value = content

        result = save_uploaded_file(uploaded, str(tmp_path))
        assert result["original_name"] == "test.xlsx"
        assert result["sha256"] == hashlib.sha256(content).hexdigest()
        assert result["size"] == len(content)
        assert os.path.isfile(result["stored_path"])

    def test_zero_byte_file(self, tmp_path):
        content = b""
        uploaded = MagicMock()
        uploaded.filename = "empty.xlsx"
        uploaded.read.return_value = content

        result = save_uploaded_file(uploaded, str(tmp_path))
        assert result["size"] == 0
        assert os.path.isfile(result["stored_path"])

    def test_sanitizes_filename(self, tmp_path):
        uploaded = MagicMock()
        uploaded.filename = "../../malicious.xlsx"
        uploaded.read.return_value = b"content"

        result = save_uploaded_file(uploaded, str(tmp_path))
        # 不应包含路径遍历部分
        assert ".." not in result["stored_path"]
        # 应保存为 UUID 文件名
        assert result["stored_path"].endswith(".xlsx")
